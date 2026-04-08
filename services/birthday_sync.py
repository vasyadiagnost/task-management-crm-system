from __future__ import annotations

import os
from datetime import datetime
from typing import Iterable

from openpyxl import Workbook, load_workbook

from models.person import Person


TEMPLATE_PATH = os.path.join("templates", "birthday_import_template.xlsx")
EXPORT_DIR = "exports"
IMPORT_SHEET_NAME = "Импорт"


def format_birthday_for_export(birthday) -> str:
    if not birthday:
        return ""

    # пока в базе у нас полная дата, поэтому отдаем полную
    return birthday.strftime("%d.%m.%Y")


def ensure_export_dir() -> None:
    os.makedirs(EXPORT_DIR, exist_ok=True)


def create_empty_template():
    wb = Workbook()
    ws = wb.active
    ws.title = IMPORT_SHEET_NAME

    # строка 1 — пример / подсказка
    ws.append([
        "Иванов Иван Иванович",
        "04.04 или 04.04.1993",
        "начальник отдела",
        "организационный отдел",
        "1",
        "+7 900 000-00-00",
        "Да",
        "сувенир",
        "можно оставить пустым — подтянется из шаблона круга",
        "поздравить лично",
    ])

    # строка 2 — реальные заголовки
    ws.append([
        "ФИО",
        "Дата рождения",
        "Должность",
        "Подразделение",
        "Круг общения",
        "Контактный телефон",
        "Поздравляем",
        "Вид подарка",
        "Текст поздравления",
        "Комментарий",
    ])

    return wb


def load_template_or_create():
    if os.path.exists(TEMPLATE_PATH):
        wb = load_workbook(TEMPLATE_PATH)
        if IMPORT_SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(IMPORT_SHEET_NAME)
            ws.append([
                "Иванов Иван Иванович",
                "04.04 или 04.04.1993",
                "начальник отдела",
                "организационный отдел",
                "1",
                "+7 900 000-00-00",
                "Да",
                "сувенир",
                "можно оставить пустым — подтянется из шаблона круга",
                "поздравить лично",
            ])
            ws.append([
                "ФИО",
                "Дата рождения",
                "Должность",
                "Подразделение",
                "Круг общения",
                "Контактный телефон",
                "Поздравляем",
                "Вид подарка",
                "Текст поздравления",
                "Комментарий",
            ])
        return wb

    return create_empty_template()


def clear_import_rows(ws):
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)


def export_persons_to_birthday_xlsx(persons: Iterable[Person]) -> str:
    ensure_export_dir()

    wb = load_template_or_create()
    ws = wb[IMPORT_SHEET_NAME]

    clear_import_rows(ws)

    for person in persons:
        ws.append([
            person.fio or "",
            format_birthday_for_export(person.birthday),
            person.position or "",
            person.department or "",
            person.circle or "",
            person.phone or "",
            "Да",          # по умолчанию
            "",            # вид подарка
            "",            # текст поздравления
            person.notes or "",
        ])

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_path = os.path.join(EXPORT_DIR, f"birthday_export_{timestamp}.xlsx")
    wb.save(output_path)

    return output_path