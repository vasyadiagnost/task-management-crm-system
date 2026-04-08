from __future__ import annotations

from datetime import datetime, date
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import inspect, text

from database import engine

SYSTEM_TABLES = {"sqlite_sequence"}

ALLOWED_STATUS_VALUES = {
    "meetings": {"Запланирована", "Пауза", "Проведена", "Отменена"},
    "tasks": {"Новая", "В работе", "Выполнена", "Отменена"},
    "registry_tasks": {"Новая", "В работе", "Выполнена", "Отменена"},
}

ALLOWED_INTERACTION_TYPES = {"Звонок", "Встреча"}

TABLE_TITLES = {
    "persons": "Шаблон миграции справочника персон",
    "interactions": "Шаблон миграции журнала контактов",
    "tasks": "Шаблон миграции задач",
    "registry_tasks": "Шаблон миграции реестра поручений",
    "meetings": "Шаблон миграции реестра регулярных встреч",
    "responsibles": "Шаблон миграции справочника ответственных",
    "circles": "Шаблон миграции справочника кругов общения",
}

COLUMN_META = {
    "persons": {
        "fio": ("ФИО", "Полное имя человека", "Текст, например: Иванов Иван Иванович"),
        "position": ("Должность", "Текущая должность", "Текст"),
        "department": ("Подразделение", "Подразделение / место работы", "Текст"),
        "circle": ("Круг общения", "Круг общения", "Текст: 1 / 2 / 3 / VIP"),
        "level": ("Уровень", "Уровень значимости / охвата", "Текст"),
        "phone": ("Телефон", "Основной телефон", "Текст"),
        "responsible": ("Ответственный", "Кто ведёт человека", "Текст из справочника ответственных"),
        "track_calls": ("Отслеживать звонки", "Нужно ли держать человека в напоминаниях", "Да / Нет"),
        "birthday": ("Дата рождения", "Дата рождения", "Дата в формате дд.мм.гггг или гггг-мм-дд"),
        "comment": ("Комментарий", "Произвольный комментарий", "Текст"),
    },
    "interactions": {
        "person_id": ("ID человека", "Ссылка на таблицу persons", "Число, например: 125"),
        "interaction_type": ("Тип контакта", "Какой тип контакта был", "Только: Звонок / Встреча"),
        "interaction_date": ("Дата контакта", "Когда состоялся контакт", "Дата в формате дд.мм.гггг или гггг-мм-дд"),
        "purpose": ("Цель контакта", "Зачем планировался контакт", "Текст"),
        "result": ("Результат", "Краткий итог общения", "Текст"),
        "next_date": ("Следующий контакт", "Когда напомнить в следующий раз", "Дата в формате дд.мм.гггг или гггг-мм-дд"),
        "responsible": ("Ответственный", "Кто ведёт контакт", "Текст"),
        "comment": ("Комментарий", "Служебная пометка", "Текст"),
        "is_active": ("Активная запись", "Текущий активный контакт или история", "0 / 1"),
        "completed_at": ("Дата завершения", "Если контакт закрыт", "Дата и время или пусто"),
    },
    "meetings": {
        "person_id": ("ID человека", "Ссылка на таблицу persons", "Число, например: 125"),
        "subject": ("Контакт / тема", "С кем или по кому планируется встреча", "Текст, обязательно"),
        "location": ("Место", "При необходимости можно указать место", "Текст или пусто"),
        "start_datetime": ("Дедлайн планирования", "Дата, к которой встречу надо поставить в календарь", "Дата дд.мм.гггг или дата-время"),
        "end_datetime": ("Окончание", "Если не используется — оставить пустым", "Дата-время или пусто"),
        "recurrence_rule": ("Периодичность", "Насколько регулярно должна быть встреча", "Например: Еженедельно / Раз в месяц"),
        "status": ("Статус", "Состояние встречи", "Только: Запланирована / Пауза / Проведена / Отменена"),
        "notes": ("Заметки", "Комментарий по встрече", "Текст"),
    },
    "tasks": {
        "person_id": ("ID человека", "Если задача привязана к человеку", "Число или пусто"),
        "title": ("Название задачи", "Краткая формулировка", "Текст, обязательно"),
        "description": ("Описание", "Расшифровка задачи", "Текст"),
        "main_responsible": ("Основной ответственный", "Кто ведёт задачу", "Текст, обязательно"),
        "co_executors": ("Соисполнители", "Кто помогает выполнять", "Текст через запятую"),
        "controller": ("Контроль", "Кто контролирует исполнение", "Текст"),
        "due_date": ("Срок", "Плановый срок", "Дата дд.мм.гггг или гггг-мм-дд"),
        "status": ("Статус", "Состояние задачи", "Только: Новая / В работе / Выполнена / Отменена"),
    },
    "registry_tasks": {
        "title": ("Название поручения", "Краткая формулировка поручения", "Текст, обязательно"),
        "description": ("Описание", "Подробности поручения", "Текст"),
        "source": ("Источник", "Кто или что является источником поручения", "Текст"),
        "main_responsible": ("Основной ответственный", "Кто ведёт поручение", "Текст, обязательно"),
        "co_executors": ("Соисполнители", "Кто помогает выполнять", "Текст через запятую"),
        "controller": ("Контроль", "Кто контролирует исполнение", "Текст"),
        "due_date": ("Срок", "Плановый срок", "Дата дд.мм.гггг или гггг-мм-дд"),
        "status": ("Статус", "Состояние поручения", "Только: Новая / В работе / Выполнена / Отменена"),
        "comment": ("Комментарий", "Служебная пометка", "Текст"),
    },
    "responsibles": {
        "name": ("ФИО ответственного", "Как будет отображаться в системе", "Текст, обязательно"),
    },
    "circles": {
        "name": ("Название круга", "Как будет отображаться в системе", "Текст, например: 1 / 2 / 3 / VIP"),
        "contact_period_days": ("Периодичность, дней", "Через сколько дней напоминать о следующем контакте", "Число, например: 14"),
    },
}


class DatabaseEditorService:
    def get_table_names(self) -> list[str]:
        inspector = inspect(engine)
        names = [name for name in inspector.get_table_names() if name not in SYSTEM_TABLES]
        names.sort()
        return names

    def get_table_columns(self, table_name: str) -> list[dict[str, Any]]:
        inspector = inspect(engine)
        columns = inspector.get_columns(table_name)
        pk_columns = set(inspector.get_pk_constraint(table_name).get("constrained_columns") or [])
        result: list[dict[str, Any]] = []
        for col in columns:
            result.append(
                {
                    "name": col["name"],
                    "type": str(col.get("type", "")),
                    "nullable": bool(col.get("nullable", True)),
                    "default": col.get("default"),
                    "primary_key": col["name"] in pk_columns,
                }
            )
        return result

    def get_primary_key_column(self, table_name: str) -> str | None:
        inspector = inspect(engine)
        pk = inspector.get_pk_constraint(table_name).get("constrained_columns") or []
        return pk[0] if pk else None

    def list_rows(self, table_name: str, limit: int = 500) -> list[dict[str, Any]]:
        pk = self.get_primary_key_column(table_name)
        order_sql = f' ORDER BY "{pk}" DESC' if pk else ""
        sql = text(f'SELECT * FROM "{table_name}"{order_sql} LIMIT :limit')
        with engine.connect() as conn:
            rows = conn.execute(sql, {"limit": int(limit)}).mappings().all()
            return [dict(row) for row in rows]

    def search_rows(self, table_name: str, query: str, limit: int = 500) -> list[dict[str, Any]]:
        query = (query or "").strip()
        if not query:
            return self.list_rows(table_name, limit=limit)

        columns = self.get_table_columns(table_name)
        searchable = [c["name"] for c in columns]
        if not searchable:
            return []

        where_parts = [f'CAST("{name}" AS TEXT) LIKE :q' for name in searchable]
        where_sql = " OR ".join(where_parts)
        pk = self.get_primary_key_column(table_name)
        order_sql = f' ORDER BY "{pk}" DESC' if pk else ""
        sql = text(f'SELECT * FROM "{table_name}" WHERE {where_sql}{order_sql} LIMIT :limit')
        with engine.connect() as conn:
            rows = conn.execute(sql, {"q": f"%{query}%", "limit": int(limit)}).mappings().all()
            return [dict(row) for row in rows]

    def update_cell(self, table_name: str, row_id: Any, column_name: str, value: Any) -> None:
        pk = self.get_primary_key_column(table_name)
        if not pk:
            raise ValueError("У таблицы нет первичного ключа")
        if column_name == pk:
            raise ValueError("Первичный ключ нельзя редактировать через Database Editor")

        typed_value = self._coerce_value(table_name, column_name, value)
        self._validate_payload(table_name, {column_name: typed_value}, partial=True)
        sql = text(f'UPDATE "{table_name}" SET "{column_name}" = :value WHERE "{pk}" = :row_id')
        with engine.begin() as conn:
            conn.execute(sql, {"value": typed_value, "row_id": row_id})

    def delete_row(self, table_name: str, row_id: Any) -> None:
        pk = self.get_primary_key_column(table_name)
        if not pk:
            raise ValueError("У таблицы нет первичного ключа")
        sql = text(f'DELETE FROM "{table_name}" WHERE "{pk}" = :row_id')
        with engine.begin() as conn:
            conn.execute(sql, {"row_id": row_id})

    def insert_row(self, table_name: str, values: dict[str, Any]) -> None:
        columns = self.get_table_columns(table_name)
        pk = self.get_primary_key_column(table_name)
        insertable = []
        params = {}
        for col in columns:
            name = col["name"]
            if name == pk and values.get(name) in (None, ""):
                continue
            if name not in values:
                continue
            insertable.append(name)
            params[name] = self._coerce_value(table_name, name, values.get(name))

        if not insertable:
            raise ValueError("Нет данных для добавления строки")

        self._validate_payload(table_name, params, partial=False)

        columns_sql = ", ".join(f'"{name}"' for name in insertable)
        values_sql = ", ".join(f':{name}' for name in insertable)
        sql = text(f'INSERT INTO "{table_name}" ({columns_sql}) VALUES ({values_sql})')
        with engine.begin() as conn:
            conn.execute(sql, params)

    def export_table_to_excel(self, table_name: str, file_path: str | Path, query: str = "") -> Path:
        rows = self.search_rows(table_name, query=query) if query else self.list_rows(table_name)
        columns = self.get_table_columns(table_name)
        wb = Workbook()
        ws = wb.active
        ws.title = table_name

        self._build_pretty_sheet(ws, table_name, columns, include_readme=False)
        start_row = 4
        technical_headers = [c["name"] for c in columns]
        for row in rows:
            values = [self._excel_output_value(row.get(col)) for col in technical_headers]
            ws.append(values)

        self._apply_data_area_formatting(ws, table_name, columns, data_start_row=start_row)
        output = Path(file_path)
        wb.save(output)
        return output

    def create_excel_template(self, table_name: str, file_path: str | Path) -> Path:
        columns = self.get_table_columns(table_name)
        wb = Workbook()
        ws = wb.active
        ws.title = "template"

        self._build_pretty_sheet(ws, table_name, columns, include_readme=True, workbook=wb)
        self._apply_data_area_formatting(ws, table_name, columns, data_start_row=4)

        output = Path(file_path)
        wb.save(output)
        return output

    def import_from_excel(self, table_name: str, file_path: str | Path, mode: str = "append") -> dict[str, Any]:
        file_path = Path(file_path)
        if not file_path.exists():
            raise ValueError("Файл Excel не найден")

        wb = load_workbook(file_path, data_only=True)
        ws = wb["template"] if "template" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("В Excel-файле нет данных")

        headers = [str(x).strip() if x is not None else "" for x in rows[0]]
        if not any(headers):
            raise ValueError("Первая строка Excel должна содержать заголовки столбцов")

        start_row_idx = 1
        while start_row_idx < len(rows) and self._looks_like_metadata_row(rows[start_row_idx]):
            start_row_idx += 1

        table_columns = self.get_table_columns(table_name)
        table_names = {c["name"] for c in table_columns}
        pk = self.get_primary_key_column(table_name)

        matched_headers = [h for h in headers if h in table_names]
        if not matched_headers:
            raise ValueError("Ни один столбец Excel не совпал со столбцами таблицы")

        if mode == "replace":
            with engine.begin() as conn:
                conn.execute(text(f'DELETE FROM "{table_name}"'))

        inserted = 0
        skipped = 0
        errors: list[str] = []

        for excel_row_index, values in enumerate(rows[start_row_idx:], start=start_row_idx + 1):
            if values is None or not any(v not in (None, "") for v in values):
                continue

            payload: dict[str, Any] = {}
            for idx, header in enumerate(headers):
                if not header or header not in table_names:
                    continue
                cell_value = values[idx] if idx < len(values) else None
                if header == pk and cell_value in (None, ""):
                    continue
                payload[header] = self._normalize_excel_value(cell_value)

            if not payload:
                skipped += 1
                continue

            try:
                self.insert_row(table_name, payload)
                inserted += 1
            except Exception as exc:
                skipped += 1
                errors.append(f"Строка {excel_row_index}: {exc}")

        return {
            "inserted": inserted,
            "skipped": skipped,
            "matched_headers": matched_headers,
            "errors": errors[:30],
        }

    def _build_pretty_sheet(self, ws, table_name: str, columns: list[dict[str, Any]], include_readme: bool, workbook=None):
        title = TABLE_TITLES.get(table_name, f"Шаблон миграции таблицы {table_name}")
        ws.freeze_panes = "A4"
        ws.sheet_view.showGridLines = False

        technical_headers = [c["name"] for c in columns]
        human_headers = [self._human_header(table_name, c["name"]) for c in columns]
        hints = [self._human_hint(table_name, c) for c in columns]

        ws.append(technical_headers)
        ws.append(human_headers)
        ws.append(hints)

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        human_fill = PatternFill("solid", fgColor="EEF4F8")
        hint_fill = PatternFill("solid", fgColor="FFF7E6")
        thin = Side(style="thin", color="C7CDD4")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, _ in enumerate(columns, start=1):
            c1 = ws.cell(row=1, column=col_idx)
            c2 = ws.cell(row=2, column=col_idx)
            c3 = ws.cell(row=3, column=col_idx)

            c1.font = Font(bold=True, color="000000")
            c1.fill = header_fill
            c1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c1.border = border

            c2.font = Font(bold=True, color="1F2937")
            c2.fill = human_fill
            c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c2.border = border

            c3.font = Font(italic=True, color="7C4A03")
            c3.fill = hint_fill
            c3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c3.border = border

            width = self._suggest_width(table_name, columns[col_idx - 1]["name"])
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 34
        ws.row_dimensions[3].height = 46
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

        if include_readme and workbook is not None:
            readme = workbook.create_sheet("README")
            readme.sheet_view.showGridLines = False
            readme.column_dimensions["A"].width = 22
            readme.column_dimensions["B"].width = 120
            readme["A1"] = "Таблица"
            readme["B1"] = table_name
            readme["A2"] = "Назначение"
            readme["B2"] = title
            readme["A4"] = "Правила"
            readme["B4"] = "1. Не менять 1 строку с техническими заголовками."
            readme["B5"] = "2. Заполнять данные начиная с 4 строки листа template."
            readme["B6"] = "3. 2 строка — человеческие названия столбцов."
            readme["B7"] = "4. 3 строка — подсказки по формату и допустимым значениям."
            readme["B8"] = "5. Массовая загрузка использует только совпадающие столбцы."
            readme["B9"] = "6. Для meetings, interactions, tasks и registry_tasks включена валидация данных."
            for cell in ["A1", "A2", "A4"]:
                readme[cell].font = Font(bold=True)
            for row in range(1, 10):
                readme[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")

    def _apply_data_area_formatting(self, ws, table_name: str, columns: list[dict[str, Any]], data_start_row: int):
        for idx, col in enumerate(columns, start=1):
            name = col["name"]
            col_letter = get_column_letter(idx)
            align = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if self._is_centered_column(name):
                align = Alignment(horizontal="center", vertical="top", wrap_text=True)
            for row in range(data_start_row, max(data_start_row + 200, ws.max_row + 1)):
                ws[f"{col_letter}{row}"].alignment = align
            if self._is_date_like_column(name):
                for row in range(data_start_row, max(data_start_row + 200, ws.max_row + 1)):
                    ws[f"{col_letter}{row}"].number_format = "DD.MM.YYYY"

    def _looks_like_metadata_row(self, row_values) -> bool:
        joined = " | ".join("" if v is None else str(v) for v in row_values).lower()
        tokens = ["тип:", "формат:", "пример:", "обязательно", "можно пусто", "полное имя", "ссылка на таблицу"]
        return any(token in joined for token in tokens)

    def _human_header(self, table_name: str, column_name: str) -> str:
        meta = COLUMN_META.get(table_name, {}).get(column_name)
        if meta:
            return meta[0]
        return column_name.replace("_", " ").capitalize()

    def _human_hint(self, table_name: str, column_meta: dict[str, Any]) -> str:
        name = column_meta["name"]
        meta = COLUMN_META.get(table_name, {}).get(name)
        if meta:
            nullable = "Можно пусто" if column_meta["nullable"] else "Обязательно"
            return f"{meta[1]}. Формат: {meta[2]}. {nullable}."
        base = f"Тип в БД: {column_meta['type']}"
        if column_meta["primary_key"]:
            base += ". Обычно не заполняется вручную"
        else:
            base += ". Заполняется при необходимости"
        if not column_meta["nullable"]:
            base += ". Поле обязательное"
        return base

    def _normalize_excel_value(self, value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value

    def _excel_output_value(self, value: Any) -> Any:
        if isinstance(value, datetime):
            return value.strftime("%d.%m.%Y %H:%M")
        if isinstance(value, date):
            return value.strftime("%d.%m.%Y")
        return value

    def _suggest_width(self, table_name: str, column_name: str) -> float:
        wide = {"fio", "title", "description", "notes", "comment", "purpose", "result", "subject", "source", "position", "department"}
        medium = {"responsible", "main_responsible", "co_executors", "controller", "recurrence_rule", "location"}
        date_cols = {"birthday", "interaction_date", "next_date", "due_date", "start_datetime", "end_datetime", "completed_at", "created_at", "updated_at"}
        if column_name in wide:
            return 30
        if column_name in medium:
            return 22
        if column_name in date_cols:
            return 18
        if column_name.endswith("_id") or column_name == "id":
            return 12
        if column_name == "status":
            return 18
        return 16

    def _is_centered_column(self, column_name: str) -> bool:
        return column_name.endswith("_id") or column_name in {"id", "status", "interaction_type", "track_calls", "is_active"}

    def _is_date_like_column(self, column_name: str) -> bool:
        return column_name in {"birthday", "interaction_date", "next_date", "due_date"}

    def _coerce_value(self, table_name: str, column_name: str, value: Any) -> Any:
        if isinstance(value, str):
            raw = value.strip()
            if raw == "":
                return None
        else:
            raw = value

        columns = {c["name"]: c for c in self.get_table_columns(table_name)}
        col = columns.get(column_name)
        type_str = (col or {}).get("type", "").upper()

        if raw is None:
            return None
        if "INT" in type_str:
            try:
                return int(raw)
            except Exception:
                return raw
        if any(token in type_str for token in ["REAL", "FLOAT", "NUMERIC", "DECIMAL", "DOUBLE"]):
            try:
                return float(str(raw).replace(",", "."))
            except Exception:
                return raw
        return raw

    def _validate_payload(self, table_name: str, payload: dict[str, Any], partial: bool = False) -> None:
        if table_name == "meetings":
            self._validate_meeting(payload, partial=partial)
        elif table_name == "interactions":
            self._validate_interaction(payload, partial=partial)
        elif table_name in {"tasks", "registry_tasks"}:
            self._validate_task_like(table_name, payload, partial=partial)

    def _validate_meeting(self, payload: dict[str, Any], partial: bool) -> None:
        if not partial and not str(payload.get("subject") or "").strip():
            raise ValueError("Для meetings поле subject обязательно")
        if "person_id" in payload and payload.get("person_id") not in (None, ""):
            if not self._is_int_like(payload.get("person_id")):
                raise ValueError("Для meetings поле person_id должно быть числом")
        if "status" in payload and payload.get("status") not in (None, ""):
            if str(payload.get("status")) not in ALLOWED_STATUS_VALUES["meetings"]:
                raise ValueError("Недопустимый статус meetings")
        for field in ["start_datetime", "end_datetime"]:
            if field in payload and payload.get(field) not in (None, ""):
                if not self._is_date_like(payload.get(field), allow_datetime=True):
                    raise ValueError(f"Для meetings поле {field} должно быть датой или датой-временем")

    def _validate_interaction(self, payload: dict[str, Any], partial: bool) -> None:
        if not partial and payload.get("person_id") in (None, ""):
            raise ValueError("Для interactions поле person_id обязательно")
        if "person_id" in payload and payload.get("person_id") not in (None, ""):
            if not self._is_int_like(payload.get("person_id")):
                raise ValueError("Для interactions поле person_id должно быть числом")
        if "interaction_type" in payload and payload.get("interaction_type") not in (None, ""):
            if str(payload.get("interaction_type")) not in ALLOWED_INTERACTION_TYPES:
                raise ValueError("interaction_type должен быть Звонок или Встреча")
        for field in ["interaction_date", "next_date", "completed_at"]:
            if field in payload and payload.get(field) not in (None, ""):
                if not self._is_date_like(payload.get(field), allow_datetime=True):
                    raise ValueError(f"Для interactions поле {field} должно быть датой")

    def _validate_task_like(self, table_name: str, payload: dict[str, Any], partial: bool) -> None:
        if not partial and not str(payload.get("title") or "").strip():
            raise ValueError(f"Для {table_name} поле title обязательно")
        if "status" in payload and payload.get("status") not in (None, ""):
            if str(payload.get("status")) not in ALLOWED_STATUS_VALUES[table_name]:
                raise ValueError(f"Недопустимый статус для {table_name}")
        if "due_date" in payload and payload.get("due_date") not in (None, ""):
            if not self._is_date_like(payload.get("due_date"), allow_datetime=True):
                raise ValueError(f"Для {table_name} поле due_date должно быть датой")

    def _is_int_like(self, value: Any) -> bool:
        try:
            int(value)
            return True
        except Exception:
            return False

    def _is_date_like(self, value: Any, allow_datetime: bool = False) -> bool:
        if isinstance(value, (datetime, date)):
            return True
        text_value = str(value).strip()
        if not text_value:
            return False
        formats = ["%Y-%m-%d", "%d.%m.%Y"]
        if allow_datetime:
            formats += ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d.%m.%Y %H:%M"]
        for fmt in formats:
            try:
                datetime.strptime(text_value, fmt)
                return True
            except ValueError:
                continue
        return False
