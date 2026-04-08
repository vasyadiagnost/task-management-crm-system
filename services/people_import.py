from __future__ import annotations

import os
from datetime import datetime, timedelta, date
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel

from database import get_session
from models.person import Person
from models.interaction import Interaction
from models.reference import Circle, Responsible
from models.app_setting import AppSetting


EXPORT_DIR = "exports"


IMPORT_TEMPLATE_CONFIG = {
    "ru": {
        "sheet_title": "Импорт людей",
        "file_name": "people_import_template.xlsx",
        "headers": [
            "ФИО",
            "Должность",
            "Круг",
            "Уровень",
            "Подкатегория",
            "Столбец 1",
            "Отслеживать звонки?",
            "пол",
            "Дата рождения",
            "Категория подарка",
            "Столбец1",
            "День рождения сегодня?",
            "Периодичность звонка",
            "Дата рождения жены",
            "Дата рождения детей",
            "Служил",
            "Проф праздник ",
            "Столбец2",
            "ПРОФ ПРАЗДНИК?",
            "Религия",
            "Хобби",
            "Телефон/Контактное лицо",
            "Аллергия/вкусы/предпочтения в еде",
            "С кем лучше не пересекать",
            "Что уже дарили ",
            "Последняя встреча ",
            "Дата следующей встречи",
            "Цель встречи ",
            "Пора встретиться?",
            "Последний звонок",
            "Дата следующего звонка",
            "Цель звонка",
            "Пора позвонить",
        ],
        "sample_row": [
            "Иванов Иван Иванович",
            "начальник отдела",
            "1",
            "регион",
            "основной контакт",
            "Шафоростов",
            "Да",
            "М",
            "04.04.1990",
            "стандарт",
            "",
            "",
            "14",
            "",
            "",
            "Нет",
            "",
            "",
            "",
            "",
            "",
            "+7 900 000-00-00",
            "",
            "",
            "",
            "01.03.2026",
            "",
            "поддержание контакта",
            "Да",
            "05.03.2026",
            "",
            "уточнить позицию",
            "Да",
        ],
    },
    "en": {
        "sheet_title": "People import",
        "file_name": "people_import_template_en.xlsx",
        "headers": [
            "Full name",
            "Position",
            "Circle",
            "Level",
            "Subcategory",
            "Column 1",
            "Track calls?",
            "Gender",
            "Birthday",
            "Gift category",
            "Column1",
            "Birthday today?",
            "Call periodicity",
            "Spouse birthday",
            "Children birthdays",
            "Served",
            "Professional holiday",
            "Column2",
            "PROFESSIONAL HOLIDAY?",
            "Religion",
            "Hobby",
            "Phone/Contact person",
            "Allergy/tastes/food preferences",
            "Who should not be crossed with",
            "What has already been gifted",
            "Last meeting",
            "Next meeting date",
            "Meeting purpose",
            "Time to meet?",
            "Last call",
            "Next call date",
            "Call purpose",
            "Time to call",
        ],
        "sample_row": [
            "Ivanov Ivan Ivanovich",
            "department head",
            "1",
            "regional",
            "primary contact",
            "Shaforostov",
            "Yes",
            "M",
            "04.04.1990",
            "standard",
            "",
            "",
            "14",
            "",
            "",
            "No",
            "",
            "",
            "",
            "",
            "",
            "+7 900 000-00-00",
            "",
            "",
            "",
            "01.03.2026",
            "",
            "relationship maintenance",
            "Yes",
            "05.03.2026",
            "",
            "clarify position",
            "Yes",
        ],
    },
}


IMPORT_HEADERS = IMPORT_TEMPLATE_CONFIG["ru"]["headers"]


HEADER_TO_ATTR = {
    "ФИО": "fio",
    "Full name": "fio",
    "Должность": "position",
    "Position": "position",
    "Круг": "circle",
    "Circle": "circle",
    "Уровень": "level",
    "Level": "level",
    "Подкатегория": "subcategory",
    "Subcategory": "subcategory",
    "Столбец 1": "legacy_column_1",
    "Column 1": "legacy_column_1",
    "Отслеживать звонки?": "track_calls",
    "Track calls?": "track_calls",
    "пол": "gender",
    "Gender": "gender",
    "Дата рождения": "birthday",
    "Birthday": "birthday",
    "Категория подарка": "gift_category",
    "Gift category": "gift_category",
    "Столбец1": "legacy_column1",
    "Column1": "legacy_column1",
    "День рождения сегодня?": "birthday_today_flag",
    "Birthday today?": "birthday_today_flag",
    "Периодичность звонка": "call_periodicity_legacy",
    "Call periodicity": "call_periodicity_legacy",
    "Дата рождения жены": "spouse_birthday",
    "Spouse birthday": "spouse_birthday",
    "Дата рождения детей": "children_birthdays",
    "Children birthdays": "children_birthdays",
    "Служил": "served",
    "Served": "served",
    "Проф праздник ": "prof_holiday",
    "Professional holiday": "prof_holiday",
    "Столбец2": "legacy_column2",
    "Column2": "legacy_column2",
    "ПРОФ ПРАЗДНИК?": "prof_holiday_flag",
    "PROFESSIONAL HOLIDAY?": "prof_holiday_flag",
    "Религия": "religion",
    "Religion": "religion",
    "Хобби": "hobby",
    "Hobby": "hobby",
    "Телефон/Контактное лицо": "phone_contact_person",
    "Phone/Contact person": "phone_contact_person",
    "Аллергия/вкусы/предпочтения в еде": "food_preferences",
    "Allergy/tastes/food preferences": "food_preferences",
    "С кем лучше не пересекать": "avoid_with",
    "Who should not be crossed with": "avoid_with",
    "Что уже дарили ": "gifts_already_given",
    "What has already been gifted": "gifts_already_given",
    "Последняя встреча ": "last_meeting_legacy",
    "Last meeting": "last_meeting_legacy",
    "Дата следующей встречи": "next_meeting_legacy",
    "Next meeting date": "next_meeting_legacy",
    "Цель встречи ": "meeting_purpose_legacy",
    "Meeting purpose": "meeting_purpose_legacy",
    "Пора встретиться?": "need_meeting_flag",
    "Time to meet?": "need_meeting_flag",
    "Последний звонок": "last_call_legacy",
    "Last call": "last_call_legacy",
    "Дата следующего звонка": "next_call_legacy",
    "Next call date": "next_call_legacy",
    "Цель звонка": "call_purpose_legacy",
    "Call purpose": "call_purpose_legacy",
    "Пора позвонить": "need_call_flag",
    "Time to call": "need_call_flag",
}


def normalize_language(language: str | None) -> str:
    return "en" if str(language or "ru").strip().lower() == "en" else "ru"


def ensure_export_dir() -> None:
    os.makedirs(EXPORT_DIR, exist_ok=True)


def create_people_import_template(*args: Any, language: str = "ru", **kwargs: Any) -> str:
    ensure_export_dir()

    if args and isinstance(args[0], str) and "language" not in kwargs:
        language = args[0]

    if "lang" in kwargs and not args:
        language = kwargs.get("lang")

    lang = normalize_language(language)
    config = IMPORT_TEMPLATE_CONFIG[lang]

    wb = Workbook()
    ws = wb.active
    ws.title = config["sheet_title"]

    ws.append(config["headers"])
    ws.append(config["sample_row"])

    output_path = os.path.join(EXPORT_DIR, config["file_name"])
    wb.save(output_path)
    return output_path


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def stringify_cell(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")

    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")

    return str(value).strip()


def normalize_yes_no(value: str) -> str:
    raw = (value or "").strip().lower()
    if raw in {"да", "yes", "y", "true", "1"}:
        return "Да"
    if raw in {"нет", "no", "n", "false", "0"}:
        return "Нет"
    return (value or "").strip()


def is_track_calls_enabled(value: str) -> bool:
    return normalize_yes_no(value) == "Да"


def parse_excel_serial_date(value: Any):
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            if isinstance(converted, datetime):
                return converted.date()
            if isinstance(converted, date):
                return converted
        except Exception:
            return None
    return None


def parse_date_flexible(value: Any):
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    excel_serial_date = parse_excel_serial_date(value)
    if excel_serial_date:
        return excel_serial_date

    raw = str(value).strip()
    if not raw:
        return None

    date_formats = [
        "%d.%m.%Y",
        "%d.%m.%y",
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%m.%d.%Y",
        "%m.%d.%y",
    ]

    for fmt in date_formats:
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue

    return None


def get_circle_period_days(session, circle_name: str) -> int:
    if not circle_name:
        return 0

    circle = session.query(Circle).filter(Circle.name == circle_name).first()
    if not circle:
        return 0

    return circle.contact_period_days or 0


def get_meeting_equals_call_setting(session) -> bool:
    setting = session.query(AppSetting).filter(AppSetting.key == "meeting_equals_call").first()
    if not setting:
        return True
    return setting.value == "1"


def deactivate_previous_active_contacts(session, person_id: int, interaction_type: str):
    query = session.query(Interaction).filter(
        Interaction.person_id == person_id,
        Interaction.is_active == 1,
    )

    if not get_meeting_equals_call_setting(session):
        query = query.filter(Interaction.interaction_type == interaction_type)

    old_items = query.all()
    for item in old_items:
        item.is_active = 0


def create_legacy_interaction(
    session,
    person: Person,
    interaction_type: str,
    interaction_date,
    next_date,
    purpose: str,
    responsible: str,
):
    if not interaction_date:
        return False

    deactivate_previous_active_contacts(session, person.id, interaction_type)

    interaction = Interaction(
        person_id=person.id,
        interaction_type=interaction_type,
        interaction_date=interaction_date,
        next_date=next_date,
        purpose=(purpose or "").strip(),
        result="",
        responsible=(responsible or person.responsible or "").strip(),
        is_active=1,
    )
    session.add(interaction)
    return True


def import_people_from_excel(file_path: str) -> dict[str, int]:
    wb = load_workbook(file_path)
    ws = wb.active

    if ws.max_row < 2:
        return {"created": 0, "updated": 0, "skipped": 0, "interactions_created": 0}

    headers = [normalize_header(cell.value) for cell in ws[1]]
    column_map: dict[int, str] = {}

    for index, header in enumerate(headers):
        attr_name = HEADER_TO_ATTR.get(header)
        if attr_name:
            column_map[index] = attr_name

    session = get_session()

    created = 0
    updated = 0
    skipped = 0
    interactions_created = 0

    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = {
                attr: stringify_cell(row[index]) if index < len(row) else ""
                for index, attr in column_map.items()
            }

            fio = row_data.get("fio", "").strip()
            if not fio:
                skipped += 1
                continue

            person = session.query(Person).filter(Person.fio == fio).first()
            is_new = person is None
            if is_new:
                person = Person(fio=fio)
                session.add(person)
                session.flush()
                created += 1
            else:
                updated += 1

            responsible_name = row_data.get("legacy_column_1", "").strip()
            if responsible_name:
                existing_responsible = session.query(Responsible).filter(Responsible.name == responsible_name).first()
                if not existing_responsible:
                    session.add(Responsible(name=responsible_name))
                    session.flush()
                person.responsible = responsible_name

            track_calls = normalize_yes_no(row_data.get("track_calls", ""))

            person.position = row_data.get("position", "").strip()
            person.circle = row_data.get("circle", "").strip()
            person.level = row_data.get("level", "").strip()
            person.subcategory = row_data.get("subcategory", "").strip()
            person.legacy_column_1 = row_data.get("legacy_column_1", "").strip()
            person.track_calls = track_calls
            person.gender = row_data.get("gender", "").strip()
            person.birthday = parse_date_flexible(row_data.get("birthday", ""))
            person.gift_category = row_data.get("gift_category", "").strip()
            person.legacy_column1 = row_data.get("legacy_column1", "").strip()
            person.birthday_today_flag = row_data.get("birthday_today_flag", "").strip()
            person.call_periodicity_legacy = row_data.get("call_periodicity_legacy", "").strip()
            person.spouse_birthday = row_data.get("spouse_birthday", "").strip()
            person.children_birthdays = row_data.get("children_birthdays", "").strip()
            person.served = row_data.get("served", "").strip()
            person.prof_holiday = row_data.get("prof_holiday", "").strip()
            person.legacy_column2 = row_data.get("legacy_column2", "").strip()
            person.prof_holiday_flag = row_data.get("prof_holiday_flag", "").strip()
            person.religion = row_data.get("religion", "").strip()
            person.hobby = row_data.get("hobby", "").strip()
            person.phone_contact_person = row_data.get("phone_contact_person", "").strip()
            person.food_preferences = row_data.get("food_preferences", "").strip()
            person.avoid_with = row_data.get("avoid_with", "").strip()
            person.gifts_already_given = row_data.get("gifts_already_given", "").strip()
            person.last_meeting_legacy = row_data.get("last_meeting_legacy", "").strip()
            person.next_meeting_legacy = row_data.get("next_meeting_legacy", "").strip()
            person.meeting_purpose_legacy = row_data.get("meeting_purpose_legacy", "").strip()
            person.need_meeting_flag = row_data.get("need_meeting_flag", "").strip()
            person.last_call_legacy = row_data.get("last_call_legacy", "").strip()
            person.next_call_legacy = row_data.get("next_call_legacy", "").strip()
            person.call_purpose_legacy = row_data.get("call_purpose_legacy", "").strip()
            person.need_call_flag = row_data.get("need_call_flag", "").strip()

            if person.circle:
                existing_circle = session.query(Circle).filter(Circle.name == person.circle).first()
                if not existing_circle:
                    period_days = 0
                    if person.circle == "1":
                        period_days = 14
                    elif person.circle == "2":
                        period_days = 21
                    elif person.circle == "3":
                        period_days = 45
                    elif person.circle.upper() == "VIP":
                        period_days = 14
                    session.add(Circle(name=person.circle, contact_period_days=period_days))
                    session.flush()

            session.flush()

            meeting_date = parse_date_flexible(row_data.get("last_meeting_legacy", ""))
            next_meeting_date = parse_date_flexible(row_data.get("next_meeting_legacy", ""))
            meeting_purpose = row_data.get("meeting_purpose_legacy", "")

            if meeting_date:
                if not next_meeting_date:
                    period_days = get_circle_period_days(session, person.circle or "")
                    if period_days > 0:
                        next_meeting_date = meeting_date + timedelta(days=period_days)
                if create_legacy_interaction(
                    session=session,
                    person=person,
                    interaction_type="Встреча",
                    interaction_date=meeting_date,
                    next_date=next_meeting_date,
                    purpose=meeting_purpose,
                    responsible=person.responsible or responsible_name,
                ):
                    interactions_created += 1

            if is_track_calls_enabled(track_calls):
                call_date = parse_date_flexible(row_data.get("last_call_legacy", ""))
                next_call_date = parse_date_flexible(row_data.get("next_call_legacy", ""))
                call_purpose = row_data.get("call_purpose_legacy", "")

                if call_date:
                    if not next_call_date:
                        period_days = get_circle_period_days(session, person.circle or "")
                        if period_days > 0:
                            next_call_date = call_date + timedelta(days=period_days)
                    if create_legacy_interaction(
                        session=session,
                        person=person,
                        interaction_type="Звонок",
                        interaction_date=call_date,
                        next_date=next_call_date,
                        purpose=call_purpose,
                        responsible=person.responsible or responsible_name,
                    ):
                        interactions_created += 1

        session.commit()
        return {
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "interactions_created": interactions_created,
        }
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()
